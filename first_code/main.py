from openpyxl import Workbook
from datetime import datetime


# test file to put date in xlsx file
wb = Workbook()

ws = wb.active

ws['A1'] = 42
ws.append([1, 2, 3])
ws['A2'] = datetime.now()

wb.save('datetime.xlsx')
